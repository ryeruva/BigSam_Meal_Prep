import openpyxl

def dump_sheets():
    wb = openpyxl.load_workbook("spring_hike_menu_19_people.xlsx", data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        filename = sheet_name.lower().replace(" ", "_").replace("—", "—").replace("(", "").replace(")", "") + ".txt"
        print(f"Dumping {sheet_name} to {filename}...")
        
        with open(filename, "w", encoding="utf-8") as f:
            f.write(f"=== SHEET: {sheet_name} ===\n\n")
            
            # We want to format it nicely as tab-separated or aligned text
            # First, find max column widths
            col_widths = {}
            for col in ws.columns:
                max_w = 0
                for cell in col:
                    val_str = str(cell.value) if cell.value is not None else ""
                    # handle newlines
                    lines = val_str.split("\n")
                    for line in lines:
                        if len(line) > max_w:
                            max_w = len(line)
                col_widths[col[0].column] = max_w
                
            # Now print row by row
            for r in range(1, ws.max_row + 1):
                row_cells = []
                # Check if it's a merged subheader (e.g. only cell in col 1 has value and others are None)
                first_val = ws.cell(row=r, column=1).value
                is_merged_subheader = False
                if first_val is not None and isinstance(first_val, str) and (first_val.startswith("PHASE") or first_val.startswith("DAY") or " — " in first_val or "TOTAL" in first_val):
                    # check if other cells in row are empty
                    others_empty = True
                    for c in range(2, 8):
                        if ws.cell(row=r, column=c).value is not None:
                            others_empty = False
                            break
                    if others_empty:
                        is_merged_subheader = True
                        
                if is_merged_subheader:
                    f.write(f"\n--- {first_val} ---\n")
                else:
                    has_val = False
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=col_idx)
                        val_str = str(cell.value) if cell.value is not None else ""
                        # replace newlines with space for row readability
                        val_str = val_str.replace("\n", " ")
                        w = col_widths.get(col_idx, 10)
                        
                        # alignment
                        if col_idx == 1:
                            # pad left
                            row_cells.append(val_str.ljust(max(w, 15)))
                        else:
                            row_cells.append(val_str.rjust(max(w, 8)))
                        if cell.value is not None:
                            has_val = True
                            
                    if has_val:
                        f.write(" | ".join(row_cells) + "\n")
            
    print("All sheets dumped successfully!")

if __name__ == "__main__":
    dump_sheets()

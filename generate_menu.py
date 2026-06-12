import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_spring_hike_menu():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Color Palette - Sleek Navy & Teal Theme (Premium)
    NAVY = "1B365D"
    TEAL = "008080"
    LIGHT_GRAY = "F2F4F7"
    ALT_ROW = "F9FAFB"
    WHITE = "FFFFFF"
    
    # Special Highlights
    ALLERGY_BG = "FFF2CC"    # Soft yellow
    ALLERGY_TXT = "800000"   # Dark red
    ADULT_BG = "FFE5CC"      # Soft orange
    ADULT_TXT = "B35A00"     # Dark orange
    TOTAL_BG = "E6EEF8"      # Soft blue-gray for totals
    
    # Fonts
    font_title = Font(name="Segoe UI", size=16, bold=True, color=NAVY)
    font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="555555")
    font_header = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    font_subheader = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=10, italic=True)
    font_allergy = Font(name="Segoe UI", size=10, bold=True, color=ALLERGY_TXT)
    font_adult = Font(name="Segoe UI", size=10, bold=True, color=ADULT_TXT)
    
    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_teal = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
    fill_light_gray = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_alt_row = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    fill_allergy = PatternFill(start_color=ALLERGY_BG, end_color=ALLERGY_BG, fill_type="solid")
    fill_adult = PatternFill(start_color=ADULT_BG, end_color=ADULT_BG, fill_type="solid")
    fill_total = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    double_side = Side(border_style="double", color="333333")
    thick_top = Side(border_style="medium", color="1B365D")
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_section_total = Border(top=thin_side, bottom=double_side)
    border_grand_total = Border(top=thick_top, bottom=double_side)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # --- 1. OVERVIEW SHEET ---
    ws_ov = wb.create_sheet(title="Overview")
    ws_ov.views.sheetView[0].showGridLines = True
    
    ws_ov["A1"] = "Ohlone Wilderness Trail — Crew Meal Plan  ·  19 People  ·  6 Bear Cans"
    ws_ov["A1"].font = font_title
    ws_ov["A2"] = "April 4–6, 2026  ·  Fremont → Lake Del Valle  ·  26.4 mi  ·  13 Scouts · 6 Adults  |  3 Scout cans (3 people) · 1 Scout can (4 people) · 2 Adult cans (3 people)"
    ws_ov["A2"].font = font_subtitle
    
    # Canister Structure Table
    ws_ov["A4"] = "BEAR CAN STRUCTURE"
    ws_ov["A4"].font = font_bold
    
    structure_headers = ["Can assignment", "Can 1", "Can 2", "Can 3", "Can 4", "Can 5", "Can 6", "Type", "People", "Notes"]
    for col_idx, header in enumerate(structure_headers, start=1):
        cell = ws_ov.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all
        
    structure_data = [
        ["Can assignment", "Scout Can 1", "Scout Can 2", "Scout Can 3", "Scout Can 4", "Adult Can 1", "Adult Can 2", "", "", "4 scout cans + 2 adult cans"],
        ["People per can", "3 scouts", "3 scouts", "3 scouts", "4 scouts", "3 adults", "3 adults", "", "", "Scouts: 3, 3, 3, 4 · Adults: 3, 3"],
        ["Hot drink", "Swiss Miss", "Swiss Miss", "Swiss Miss", "Swiss Miss", "own coffee", "own coffee", "", "", "Scouts: Swiss Miss · Adults: own coffee/tea"],
        ["Cooks for", "Themselves", "Themselves", "Themselves", "Themselves", "Themselves", "Themselves", "", "", "Each bear can group cooks independently"]
    ]
    
    for r_idx, row_data in enumerate(structure_data, start=6):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_ov.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if c_idx == 1 else font_regular
            cell.alignment = align_left if c_idx == 1 or c_idx == 10 else align_center
            cell.border = border_all
            if r_idx % 2 == 1:
                cell.fill = fill_alt_row
                
    # Trip Schedule Table
    ws_ov["A11"] = "TRIP SCHEDULE"
    ws_ov["A11"].font = font_bold
    
    schedule_headers = ["Day", "Date", "Route", "Miles", "Gain", "Loss", "Camp", "Meals", "Water", "Notes"]
    for col_idx, header in enumerate(schedule_headers, start=1):
        cell = ws_ov.cell(row=12, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all
        
    schedule_data = [
        ["Day 1", "Sat Apr 4", "Trailhead → Star's Rest", "10.92 mi", "+3,474 ft", "−2,272 ft", "Star's Rest", "Snacks · Lunch · Dinner · Dessert", "Creek (filter)", "Eat breakfast before leaving car"],
        ["Day 2", "Sun Apr 5", "Star's Rest → Maggie's Half Acre", "5.82 mi", "+2,316 ft", "−410 ft", "Maggie's Half Acre", "Bkfst · Snacks · Lunch · Dinner · Dessert", "Dry ridge — fill at camp", "Hardest climb day; cattle on trail"],
        ["Day 3", "Mon Apr 6", "Maggie's Half Acre → Lake Del Valle", "9.69 mi", "+1,514 ft", "−4,249 ft", "Exit — Del Valle", "Bkfst · Snacks · Lunch", "Tap at park", "Mostly downhill; celebrate at Del Valle"],
        ["TOTALS", "", "", "26.43 mi", "+7,304 ft", "−6,931 ft", "", "", "", ""]
    ]
    
    for r_idx, row_data in enumerate(schedule_data, start=13):
        is_total = (row_data[0] == "TOTALS")
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_ov.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if is_total or c_idx == 1 else font_regular
            cell.alignment = align_left if c_idx == 3 or c_idx == 10 else align_center
            cell.border = border_all
            if is_total:
                cell.fill = fill_total
                cell.border = Border(top=thin_side, bottom=double_side, left=thin_side, right=thin_side)
            elif r_idx % 2 == 1:
                cell.fill = fill_alt_row
                
    # Calorie Overview
    ws_ov["A18"] = "CALORIE OVERVIEW (per person)"
    ws_ov["A18"].font = font_bold
    
    calorie_headers = ["", "Scouts", "Adults", "Notes"]
    for col_idx, header in enumerate(calorie_headers, start=1):
        cell = ws_ov.cell(row=19, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_all
        
    calorie_data = [
        ["Target (full active day)", "~2,750 kcal", "~3,750 kcal", "Sunday is the only full active day"],
        ["Sat Apr 4 delivered", "='Meal Plan (Per Bear Can)'!F39", "='Meal Plan (Per Bear Can)'!H39", "Snacks + Lunch + Dinner + Dessert (no breakfast carried)"],
        ["Sun Apr 5 delivered", "='Meal Plan (Per Bear Can)'!F89", "='Meal Plan (Per Bear Can)'!H89", "Bkfst + Snacks (incl. adult supplement) + Lunch + Dinner + Dessert"],
        ["Mon Apr 6 delivered", "='Meal Plan (Per Bear Can)'!F114", "='Meal Plan (Per Bear Can)'!H114", "Bkfst + Snacks + Lunch"],
        ["TRIP TOTAL", "='Meal Plan (Per Bear Can)'!F118", "='Meal Plan (Per Bear Can)'!H118", "3 days combined per person"]
    ]
    
    for r_idx, row_data in enumerate(calorie_data, start=20):
        is_total = (row_data[0] == "TRIP TOTAL")
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_ov.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if is_total or c_idx == 1 else font_regular
            cell.alignment = align_left if c_idx == 1 or c_idx == 4 else align_center
            cell.border = border_all
            if is_total:
                cell.fill = fill_total
                cell.border = Border(top=thin_side, bottom=double_side, left=thin_side, right=thin_side)
                if c_idx in [2, 3]:
                    cell.number_format = '"~"#,##0" kcal"'
            elif r_idx % 2 == 1:
                cell.fill = fill_alt_row
            
            if not is_total and c_idx in [2, 3] and r_idx > 20:
                cell.number_format = '"~"#,##0" kcal"'
                
    # Budget Overview
    ws_ov["A26"] = "BUDGET & WEIGHT OVERVIEW"
    ws_ov["A26"].font = font_bold
    
    budget_headers = ["", "Adult Can (3 ppl) (×2)", "Scout Can (3 ppl) (×3)", "Scout Can (4 ppl) (×1)", "Group Total", "Notes"]
    for col_idx, header in enumerate(budget_headers, start=1):
        cell = ws_ov.cell(row=27, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_all
        
    budget_data = [
        ["Estimated food cost", "=SUM('Shopping List'!J$4:J$31)", "=SUM('Shopping List'!H$4:H$31)", "=SUM('Shopping List'!I$4:I$31)", "=SUM('Shopping List'!K$4:K$31)", "Bulk prices (Costco/Amazon)"],
        ["Per person", "=B28/3", "=C28/3", "=D28/4", "=E28/19", "Averaged across group"],
        ["Food weight (dry)", "=SUMPRODUCT('Shopping List'!L$4:L$31,'Shopping List'!E$4:E$31)/16", "=SUMPRODUCT('Shopping List'!L$4:L$31,'Shopping List'!C$4:C$31)/16", "=SUMPRODUCT('Shopping List'!L$4:L$31,'Shopping List'!D$4:D$31)/16", "=SUM('Shopping List'!M$4:M$31)/16", "Dry food weight in lbs"]
    ]
    
    for r_idx, row_data in enumerate(budget_data, start=28):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_ov.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if c_idx == 1 or c_idx == 5 else font_regular
            cell.alignment = align_left if c_idx == 1 or c_idx == 6 else align_center
            cell.border = border_all
            if r_idx % 2 == 1:
                cell.fill = fill_alt_row
                
            # Formatting
            if r_idx in [28, 29] and c_idx in [2, 3, 4, 5]:
                cell.number_format = '"$"#,##0.00'
            elif r_idx == 30 and c_idx in [2, 3, 4, 5]:
                cell.number_format = '0.0" lbs"'
                
    # Water & Safety notes (keep static text as in original)
    ws_ov["A32"] = "WATER & SAFETY INFORMATION"
    ws_ov["A32"].font = font_bold
    
    water_headers = ["Location", "Source", "Reliability", "Required Action"]
    for col_idx, header in enumerate(water_headers, start=1):
        cell = ws_ov.cell(row=33, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all
        
    water_data = [
        ["Trailhead (Fremont)", "Tap / car fill", "Reliable", "Fill ALL bottles. Carry 2-3 L per person."],
        ["Star's Rest Camp", "Seasonal creek", "Usually present", "Filter before use. Refill fully before leaving."],
        ["Day 2 Ridgeline", "None", "⚠ DRY", "No water on ridge. Must carry enough from camp."],
        ["Maggie's Half Acre", "Seasonal creek", "Usually present", "Filter before use. Refill fully before bed."],
        ["Del Valle (exit)", "Park tap", "Reliable", "Refill at the end of the trail."]
    ]
    
    for r_idx, row_data in enumerate(water_data, start=34):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_ov.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.alignment = align_left
            cell.border = border_all
            if r_idx % 2 == 1:
                cell.fill = fill_alt_row
                
    # --- 2. MEAL PLAN (PER BEAR CAN) SHEET ---
    ws_mp = wb.create_sheet(title="Meal Plan (Per Bear Can)")
    ws_mp.views.sheetView[0].showGridLines = True
    
    ws_mp["A1"] = "Meal Plan — Quantities & Nutritional Breakdown Per Bear Can Type"
    ws_mp["A1"].font = font_title
    ws_mp["A2"] = "Calculated for: Scout Can (3 people), Scout Can (4 people), and Adult Can (3 people)  |  ⚠ Highlighted = Contains Nuts (swap for allergy scout)"
    ws_mp["A2"].font = font_subtitle
    
    # Detailed meal plan layout
    # Headers: Food Item | Scout 3-ppl Qty | Scout 4-ppl Qty | Adult 3-ppl Qty | kcal/unit | Scout 3 kcal | Scout 4 kcal | Adult 3 kcal | $/unit | Scout 3 $ | Scout 4 $ | Adult 3 $ | oz/unit | Scout 3 oz | Scout 4 oz | Adult 3 oz | Notes
    headers = [
        "Food Item", "Qty\n(Scout 3)", "Qty\n(Scout 4)", "Qty\n(Adult 3)",
        "kcal/unit", "Scout 3\nkcal", "Scout 4\nkcal", "Adult 3\nkcal",
        "$/unit", "Scout 3\n$", "Scout 4\n$", "Adult 3\n$",
        "oz/unit", "Scout 3\noz", "Scout 4\noz", "Adult 3\noz", "Notes"
    ]
    
    # We will write day-by-day meal plan
    # Helper to write subheader row
    def write_subheader(ws, r, title_text, fill_color):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
        cell = ws.cell(row=r, column=1, value=title_text)
        cell.font = font_subheader
        cell.fill = fill_color
        cell.alignment = align_left
        ws.row_dimensions[r].height = 24
        
    def write_table_headers(ws, r):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=9, bold=True, color=WHITE)
            cell.fill = fill_teal
            cell.alignment = align_wrap_center
            cell.border = border_all
        ws.row_dimensions[r].height = 30
        
    # Main meal plan data
    # Saturday
    r = 4
    write_subheader(ws_mp, r, "SAT APR 4 — BREAKFAST (Eaten at home before driving to trailhead)", fill_navy)
    
    r = 5
    write_subheader(ws_mp, r, "SAT APR 4 — SNACKS (Handed out at trailhead)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sat_snacks_start = r + 1
    sat_snacks_data = [
        # name, sq3, sq4, ad3, kcal, price, oz, notes, is_nut
        ["Kind – Peanut Butter Dark Chocolate Bar", 3, 4, 3, 200, 1.04, 1.4, "⚠ NUT: swap → 88 Acres bar for allergy scout", True],
        ["Kind – Caramel Almond & Peanut Bar", 3, 4, 3, 180, 1.59, 1.4, "⚠ NUT: swap → 88 Acres bar for allergy scout", True],
        ["Krave Sea Salt Original Beef Jerky", 3, 4, 3, 140, 2.19, 1.48, "Safe for nut allergy ✓", False],
        ["Gatorade Powder Pack", 3, 4, 3, 130, 0.38, 1.23, "Electrolytes", False],
        ["Trader Joe's Dried Mango Slices (1 oz)", 3, 4, 3, 90, 0.55, 1.0, "Trader Joe's — Sat snack · nut-free ✓", False]
    ]
    
    for item in sat_snacks_data:
        r += 1
        # Write values
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        # kcal formulas
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        # price
        ws_mp.cell(row=r, column=9, value=item[5])
        # cost formulas
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        # weight
        ws_mp.cell(row=r, column=13, value=item[6])
        # weight formulas
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        # Styles
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            # alignments
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                
            # number formats
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sat_snacks_end = r
    
    # Sat Snacks Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Snacks Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sat_snacks_start}:F{sat_snacks_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sat_snacks_start}:G{sat_snacks_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sat_snacks_start}:H{sat_snacks_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sat_snacks_start}:J{sat_snacks_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sat_snacks_start}:K{sat_snacks_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sat_snacks_start}:L{sat_snacks_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sat_snacks_start}:N{sat_snacks_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sat_snacks_start}:O{sat_snacks_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sat_snacks_start}:P{sat_snacks_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
    
    sat_snacks_total_row = r
    
    # Saturday Lunch
    r += 1
    write_subheader(ws_mp, r, "SAT APR 4 — LUNCH (NO COOK — trail charcuterie)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sat_lunch_start = r + 1
    sat_lunch_data = [
        ["Columbus Hard Salami (sliced packet)", 3, 4, 3, 220, 2.99, 2.0, "1 packet/person; nut-free ✓", False],
        ["Wasa Crisp'n Light Crackerbread", 3, 4, 3, 100, 0.50, 1.0, "1 pack/person; nut-free ✓", False],
        ["Tillamook Cheddar Cheese Packets", 3, 4, 3, 90, 0.75, 0.7, "1 packet/person; nut-free ✓", False]
    ]
    
    for item in sat_lunch_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None)
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font_regular
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sat_lunch_end = r
    
    # Sat Lunch Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Lunch Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sat_lunch_start}:F{sat_lunch_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sat_lunch_start}:G{sat_lunch_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sat_lunch_start}:H{sat_lunch_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sat_lunch_start}:J{sat_lunch_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sat_lunch_start}:K{sat_lunch_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sat_lunch_start}:L{sat_lunch_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sat_lunch_start}:N{sat_lunch_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sat_lunch_start}:O{sat_lunch_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sat_lunch_start}:P{sat_lunch_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sat_lunch_total_row = r
    
    # Saturday Dinner
    r += 1
    write_subheader(ws_mp, r, "SAT APR 4 — DINNER (hot, communal — Star's Rest)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sat_dinner_start = r + 1
    sat_dinner_data = [
        # Peak Refuel Pasta kcal/unit corrected to 780 (per pouch, which is 1.5 svgs of 520 kcal each)
        ["Peak Refuel – Beef Pasta Marinara", 2, 3, 3, 780, 7.49, 3.2, "Scout 3: 2 pouches · Scout 4: 3 pouches · Adult 3: 3 pouches", False],
        ["Mission Street Taco Flour Tortillas", 6, 8, 6, 75, 0.25, 0.5, "2/person; nut-free ✓", False],
        ["Tillamook Cheddar Cheese Packets", 6, 8, 6, 90, 0.75, 0.7, "2/person; nut-free ✓", False],
        ["Swiss Miss Hot Chocolate", 6, 8, 0, 90, 0.38, 0.73, "Scout cans only (2/scout); Adults: own coffee/tea", False]
    ]
    
    for item in sat_dinner_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None)
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font_regular
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sat_dinner_end = r
    
    # Sat Dinner Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Dinner Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sat_dinner_start}:F{sat_dinner_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sat_dinner_start}:G{sat_dinner_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sat_dinner_start}:H{sat_dinner_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sat_dinner_start}:J{sat_dinner_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sat_dinner_start}:K{sat_dinner_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sat_dinner_start}:L{sat_dinner_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sat_dinner_start}:N{sat_dinner_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sat_dinner_start}:O{sat_dinner_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sat_dinner_start}:P{sat_dinner_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sat_dinner_total_row = r
    
    # Saturday Dessert
    r += 1
    write_subheader(ws_mp, r, "SAT APR 4 — DESSERT (Crème Brûlée)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sat_dessert_start = r + 1
    # BP Crème Brûlée kcal/unit corrected to 620 (per pouch, which is 2 servings of 310 kcal each)
    sat_dessert_data = [
        ["Backpacker's Pantry – Crème Brûlée", 2, 2, 2, 620, 8.99, 2.26, "1 pouch per 2 people; buddy pairs share. ⚠ Check label for cashews", False]
    ]
    
    for item in sat_dessert_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None)
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font_regular
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sat_dessert_end = r
    
    # Sat Dessert Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Dessert Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sat_dessert_start}:F{sat_dessert_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sat_dessert_start}:G{sat_dessert_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sat_dessert_start}:H{sat_dessert_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sat_dessert_start}:J{sat_dessert_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sat_dessert_start}:K{sat_dessert_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sat_dessert_start}:L{sat_dessert_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sat_dessert_start}:N{sat_dessert_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sat_dessert_start}:O{sat_dessert_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sat_dessert_start}:P{sat_dessert_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sat_dessert_total_row = r
    
    # SATURDAY DAILY TOTAL (PER PERSON CALORIES & TOTAL CAN VALUE)
    r += 1
    ws_mp.cell(row=r, column=1, value="SAT APR 4 — DAILY TOTAL (Per Bear Can)")
    ws_mp.cell(row=r, column=6, value=f"=F{sat_snacks_total_row}+F{sat_lunch_total_row}+F{sat_dinner_total_row}+F{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=7, value=f"=G{sat_snacks_total_row}+G{sat_lunch_total_row}+G{sat_dinner_total_row}+G{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=8, value=f"=H{sat_snacks_total_row}+H{sat_lunch_total_row}+H{sat_dinner_total_row}+H{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=10, value=f"=J{sat_snacks_total_row}+J{sat_lunch_total_row}+J{sat_dinner_total_row}+J{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=11, value=f"=K{sat_snacks_total_row}+K{sat_lunch_total_row}+K{sat_dinner_total_row}+K{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=12, value=f"=L{sat_snacks_total_row}+L{sat_lunch_total_row}+L{sat_dinner_total_row}+L{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=14, value=f"=N{sat_snacks_total_row}+N{sat_lunch_total_row}+N{sat_dinner_total_row}+N{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=15, value=f"=O{sat_snacks_total_row}+O{sat_lunch_total_row}+O{sat_dinner_total_row}+O{sat_dessert_total_row}")
    ws_mp.cell(row=r, column=16, value=f"=P{sat_snacks_total_row}+P{sat_lunch_total_row}+P{sat_dinner_total_row}+P{sat_dessert_total_row}")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_navy
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.border = border_grand_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sat_daily_total_row = r
    
    # SATURDAY PER PERSON
    r += 1
    ws_mp.cell(row=r, column=1, value="SAT APR 4 — CALORIES DELIVERED PER PERSON")
    ws_mp.cell(row=r, column=6, value=f"=F{sat_daily_total_row}/3") # Scout 3
    ws_mp.cell(row=r, column=7, value=f"=G{sat_daily_total_row}/4") # Scout 4
    ws_mp.cell(row=r, column=8, value=f"=H{sat_daily_total_row}/3") # Adult 3
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = Border(top=thin_side, bottom=thick_top)
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
            
    sat_per_person_row = r
    
    # --- SUNDAY ---
    r += 1
    write_subheader(ws_mp, r, "SUN APR 5 — BREAKFAST (grab & go — MET-Rx Big100)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sun_bkfst_start = r + 1
    sun_bkfst_data = [
        ["MET-Rx Big100 – Super Cookie Crunch", 3, 4, 3, 420, 2.19, 3.53, "1/person; ⚠ NUT: swap → No Nuts! bar for allergy scout", True],
        ["Swiss Miss Hot Chocolate", 6, 8, 0, 90, 0.38, 0.73, "Scout cans only — boil water for drinks ONLY", False]
    ]
    
    for item in sun_bkfst_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sun_bkfst_end = r
    
    # Sun Bkfst Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Breakfast Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sun_bkfst_start}:F{sun_bkfst_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sun_bkfst_start}:G{sun_bkfst_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sun_bkfst_start}:H{sun_bkfst_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sun_bkfst_start}:J{sun_bkfst_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sun_bkfst_start}:K{sun_bkfst_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sun_bkfst_start}:L{sun_bkfst_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sun_bkfst_start}:N{sun_bkfst_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sun_bkfst_start}:O{sun_bkfst_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sun_bkfst_start}:P{sun_bkfst_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_bkfst_total_row = r
    
    # Sunday Snacks
    r += 1
    write_subheader(ws_mp, r, "SUN APR 5 — SNACKS (individual — all cans)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sun_snacks_start = r + 1
    sun_snacks_data = [
        ["Krave Sea Salt Original Beef Jerky", 3, 4, 3, 140, 2.19, 1.48, "Safe ✓", False],
        ["Kind – Dark Choc Nuts & Sea Salt Bar", 3, 4, 3, 200, 1.04, 1.4, "⚠ NUT: swap → 88 Acres bar for allergy scout", True],
        ["Gatorade Powder Pack", 6, 8, 6, 130, 0.38, 1.23, "×2/person; hardest climbing day", False],
        ["Banana Chips (1 oz/person)", 3, 4, 3, 150, 0.45, 1.0, "nut-free ✓", False]
    ]
    
    for item in sun_snacks_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sun_snacks_end = r
    
    # Sun Snacks Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Snacks Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sun_snacks_start}:F{sun_snacks_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sun_snacks_start}:G{sun_snacks_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sun_snacks_start}:H{sun_snacks_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sun_snacks_start}:J{sun_snacks_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sun_snacks_start}:K{sun_snacks_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sun_snacks_start}:L{sun_snacks_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sun_snacks_start}:N{sun_snacks_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sun_snacks_start}:O{sun_snacks_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sun_snacks_start}:P{sun_snacks_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_snacks_total_row = r
    
    # Sunday Adult Supplement
    r += 1
    write_subheader(ws_mp, r, "SUN APR 5 — ADULT SUPPLEMENT (Adult Cans Only)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sun_supp_start = r + 1
    sun_supp_data = [
        ["Clif Builder's – Crunchy Peanut Butter", 0, 0, 3, 280, 2.17, 2.4, "Pre-dinner — adult cans only; 1/person; ⚠ NUT", True],
        ["Justin's Classic PB Squeeze Pack", 0, 0, 3, 210, 1.49, 1.15, "Any time — adult cans only; 1/person; ⚠ NUT", True]
    ]
    
    for item in sun_supp_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_adult
        font = font_adult
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sun_supp_end = r
    
    # Sun Supp Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Adult Supplement Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sun_supp_start}:F{sun_supp_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sun_supp_start}:G{sun_supp_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sun_supp_start}:H{sun_supp_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sun_supp_start}:J{sun_supp_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sun_supp_start}:K{sun_supp_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sun_supp_start}:L{sun_supp_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sun_supp_start}:N{sun_supp_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sun_supp_start}:O{sun_supp_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sun_supp_start}:P{sun_supp_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_supp_total_row = r
    
    # Sunday Lunch
    r += 1
    ws_mp.cell(row=r, column=1, value="SUN APR 5 — LUNCH (NO COOK — chicken wraps)")
    ws_mp.cell(row=r, column=1).font = font_subheader
    ws_mp.cell(row=r, column=1).fill = fill_teal
    ws_mp.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
    ws_mp.row_dimensions[r].height = 24
    r += 1
    write_table_headers(ws_mp, r)
    
    sun_lunch_start = r + 1
    sun_lunch_data = [
        ["Chicken of the Sea Shredded Chicken Pouch", 3, 4, 3, 150, 2.29, 2.5, "1/person; nut-free ✓", False],
        ["Mission Street Taco Flour Tortillas", 6, 8, 6, 75, 0.25, 0.5, "2/person; nut-free ✓", False],
        ["Tillamook Cheddar Cheese Packets", 3, 4, 3, 90, 0.75, 0.7, "1/person; nut-free ✓", False],
        ["Clif Bar – Crunchy Peanut Butter", 3, 4, 3, 250, 1.29, 2.4, "⚠ NUT: swap → Clif Chocolate Chip for allergy scout", True],
        ["Justin's Classic PB Squeeze Pack", 3, 4, 3, 210, 1.49, 1.15, "⚠ NUT: swap → Sunflower Butter Squeeze for allergy scout", True]
    ]
    
    for item in sun_lunch_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sun_lunch_end = r
    
    # Sun Lunch Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Lunch Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sun_lunch_start}:F{sun_lunch_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sun_lunch_start}:G{sun_lunch_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sun_lunch_start}:H{sun_lunch_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sun_lunch_start}:J{sun_lunch_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sun_lunch_start}:K{sun_lunch_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sun_lunch_start}:L{sun_lunch_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sun_lunch_start}:N{sun_lunch_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sun_lunch_start}:O{sun_lunch_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sun_lunch_start}:P{sun_lunch_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_lunch_total_row = r
    
    # Sunday Dinner
    r += 1
    write_subheader(ws_mp, r, "SUN APR 5 — DINNER (hot, communal — Maggie's Half Acre)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sun_dinner_start = r + 1
    # Peak Stroganoff corrected to 600 kcal/unit (per pouch, which is 1.5 svgs of 400 kcal each)
    sun_dinner_data = [
        ["Peak Refuel – Beef Stroganoff", 2, 3, 3, 600, 7.49, 3.5, "Scout 3: 2 pouches · Scout 4: 3 pouches · Adult 3: 3 pouches", False],
        ["Mission Street Taco Flour Tortillas", 6, 8, 6, 75, 0.25, 0.5, "2/person; nut-free ✓", False],
        ["Idahoan Mashed Potatoes – Butter & Herb", 3, 4, 3, 110, 0.98, 1.0, "1/person; nut-free ✓", False],
        ["Swiss Miss Hot Chocolate", 6, 8, 0, 90, 0.38, 0.73, "Scout cans only (2/scout); Adults: own coffee/tea", False]
    ]
    
    for item in sun_dinner_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None)
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font_regular
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sun_dinner_end = r
    
    # Sun Dinner Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Dinner Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sun_dinner_start}:F{sun_dinner_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sun_dinner_start}:G{sun_dinner_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sun_dinner_start}:H{sun_dinner_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sun_dinner_start}:J{sun_dinner_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sun_dinner_start}:K{sun_dinner_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sun_dinner_start}:L{sun_dinner_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sun_dinner_start}:N{sun_dinner_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sun_dinner_start}:O{sun_dinner_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sun_dinner_start}:P{sun_dinner_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_dinner_total_row = r
    
    # Sunday Dessert
    r += 1
    write_subheader(ws_mp, r, "SUN APR 5 — DESSERT (Justin's Dark Choc PB Cups)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    sun_dessert_start = r + 1
    sun_dessert_data = [
        ["Justin's Dark Chocolate Peanut Butter Cups", 3, 4, 3, 220, 1.49, 1.41, "1/person; ⚠ NUT: swap → Honey Stinger waffle for allergy scout", True]
    ]
    
    for item in sun_dessert_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    sun_dessert_end = r
    
    # Sun Dessert Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Dessert Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{sun_dessert_start}:F{sun_dessert_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{sun_dessert_start}:G{sun_dessert_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{sun_dessert_start}:H{sun_dessert_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{sun_dessert_start}:J{sun_dessert_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{sun_dessert_start}:K{sun_dessert_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{sun_dessert_start}:L{sun_dessert_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{sun_dessert_start}:N{sun_dessert_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{sun_dessert_start}:O{sun_dessert_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{sun_dessert_start}:P{sun_dessert_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_dessert_total_row = r
    
    # SUNDAY DAILY TOTAL
    r += 1
    ws_mp.cell(row=r, column=1, value="SUN APR 5 — DAILY TOTAL (Per Bear Can)")
    ws_mp.cell(row=r, column=6, value=f"=F{sun_bkfst_total_row}+F{sun_snacks_total_row}+F{sun_supp_total_row}+F{sun_lunch_total_row}+F{sun_dinner_total_row}+F{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=7, value=f"=G{sun_bkfst_total_row}+G{sun_snacks_total_row}+G{sun_supp_total_row}+G{sun_lunch_total_row}+G{sun_dinner_total_row}+G{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=8, value=f"=H{sun_bkfst_total_row}+H{sun_snacks_total_row}+H{sun_supp_total_row}+H{sun_lunch_total_row}+H{sun_dinner_total_row}+H{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=10, value=f"=J{sun_bkfst_total_row}+J{sun_snacks_total_row}+J{sun_supp_total_row}+J{sun_lunch_total_row}+J{sun_dinner_total_row}+J{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=11, value=f"=K{sun_bkfst_total_row}+K{sun_snacks_total_row}+K{sun_supp_total_row}+K{sun_lunch_total_row}+K{sun_dinner_total_row}+K{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=12, value=f"=L{sun_bkfst_total_row}+L{sun_snacks_total_row}+L{sun_supp_total_row}+L{sun_lunch_total_row}+L{sun_dinner_total_row}+L{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=14, value=f"=N{sun_bkfst_total_row}+N{sun_snacks_total_row}+N{sun_supp_total_row}+N{sun_lunch_total_row}+N{sun_dinner_total_row}+N{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=15, value=f"=O{sun_bkfst_total_row}+O{sun_snacks_total_row}+O{sun_supp_total_row}+O{sun_lunch_total_row}+O{sun_dinner_total_row}+O{sun_dessert_total_row}")
    ws_mp.cell(row=r, column=16, value=f"=P{sun_bkfst_total_row}+P{sun_snacks_total_row}+P{sun_supp_total_row}+P{sun_lunch_total_row}+P{sun_dinner_total_row}+P{sun_dessert_total_row}")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_navy
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.border = border_grand_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    sun_daily_total_row = r
    
    # SUNDAY PER PERSON
    r += 1
    ws_mp.cell(row=r, column=1, value="SUN APR 5 — CALORIES DELIVERED PER PERSON")
    ws_mp.cell(row=r, column=6, value=f"=F{sun_daily_total_row}/3") # Scout 3
    ws_mp.cell(row=r, column=7, value=f"=G{sun_daily_total_row}/4") # Scout 4
    ws_mp.cell(row=r, column=8, value=f"=H{sun_daily_total_row}/3") # Adult 3
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = Border(top=thin_side, bottom=thick_top)
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
            
    sun_per_person_row = r
    
    # --- MONDAY ---
    r += 1
    write_subheader(ws_mp, r, "MON APR 6 — BREAKFAST (grab & go — ProBar Meal)", fill_teal)
    r += 1
    write_table_headers(ws_mp, r)
    
    mon_bkfst_start = r + 1
    mon_bkfst_data = [
        ["ProBar Meal – Oatmeal Chocolate Chip", 3, 4, 3, 410, 2.99, 3.0, "1/person; ⚠ NUT (cashews): swap → No Nuts! bar for allergy scout", True],
        ["Swiss Miss Hot Chocolate", 6, 8, 0, 90, 0.38, 0.73, "Scout cans only — boil water for drinks ONLY", False]
    ]
    
    for item in mon_bkfst_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    mon_bkfst_end = r
    
    # Mon Bkfst Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Breakfast Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{mon_bkfst_start}:F{mon_bkfst_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{mon_bkfst_start}:G{mon_bkfst_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{mon_bkfst_start}:H{mon_bkfst_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{mon_bkfst_start}:J{mon_bkfst_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{mon_bkfst_start}:K{mon_bkfst_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{mon_bkfst_start}:L{mon_bkfst_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{mon_bkfst_start}:N{mon_bkfst_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{mon_bkfst_start}:O{mon_bkfst_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{mon_bkfst_start}:P{mon_bkfst_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    mon_bkfst_total_row = r
    
    # Monday Snacks
    r += 1
    ws_mp.cell(row=r, column=1, value="MON APR 6 — SNACKS (hip-belt pocket)")
    ws_mp.cell(row=r, column=1).font = font_subheader
    ws_mp.cell(row=r, column=1).fill = fill_teal
    ws_mp.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
    ws_mp.row_dimensions[r].height = 24
    r += 1
    write_table_headers(ws_mp, r)
    
    mon_snacks_start = r + 1
    mon_snacks_data = [
        ["Clif Bar – Crunchy Peanut Butter", 3, 4, 3, 250, 1.29, 2.4, "⚠ NUT: swap → Clif Chocolate Chip for allergy scout", True],
        ["Krave Sea Salt Original Beef Jerky", 3, 4, 3, 140, 2.19, 1.48, "Safe ✓", False],
        ["Gatorade Powder Pack", 3, 4, 3, 130, 0.38, 1.23, "×1/person", False],
        # Brownie bites: 2 pouches per canister type
        ["Peak Refuel – Fudge Brownie Bites", 2, 2, 2, 300, 4.99, 1.5, "Mon afternoon snack · no cook · nut-free ✓", False]
    ]
    
    for item in mon_snacks_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    mon_snacks_end = r
    
    # Mon Snacks Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Snacks Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{mon_snacks_start}:F{mon_snacks_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{mon_snacks_start}:G{mon_snacks_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{mon_snacks_start}:H{mon_snacks_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{mon_snacks_start}:J{mon_snacks_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{mon_snacks_start}:K{mon_snacks_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{mon_snacks_start}:L{mon_snacks_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{mon_snacks_start}:N{mon_snacks_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{mon_snacks_start}:O{mon_snacks_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{mon_snacks_start}:P{mon_snacks_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    mon_snacks_total_row = r
    
    # Monday Lunch
    r += 1
    ws_mp.cell(row=r, column=1, value="MON APR 6 — LUNCH (NO COOK — PB wraps)")
    ws_mp.cell(row=r, column=1).font = font_subheader
    ws_mp.cell(row=r, column=1).fill = fill_teal
    ws_mp.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
    ws_mp.row_dimensions[r].height = 24
    r += 1
    write_table_headers(ws_mp, r)
    
    mon_lunch_start = r + 1
    mon_lunch_data = [
        ["Mission Street Taco Flour Tortillas", 6, 8, 6, 75, 0.25, 0.5, "2/person; nut-free ✓", False],
        ["Justin's Classic PB Squeeze Pack", 3, 4, 3, 190, 1.49, 1.15, "1/person; ⚠ NUT: swap → Sunflower Butter Squeeze for allergy scout", True]
    ]
    
    for item in mon_lunch_data:
        r += 1
        ws_mp.cell(row=r, column=1, value=item[0])
        ws_mp.cell(row=r, column=2, value=item[1])
        ws_mp.cell(row=r, column=3, value=item[2])
        ws_mp.cell(row=r, column=4, value=item[3])
        ws_mp.cell(row=r, column=5, value=item[4])
        ws_mp.cell(row=r, column=6, value=f"=B{r}*E{r}")
        ws_mp.cell(row=r, column=7, value=f"=C{r}*E{r}")
        ws_mp.cell(row=r, column=8, value=f"=D{r}*E{r}")
        ws_mp.cell(row=r, column=9, value=item[5])
        ws_mp.cell(row=r, column=10, value=f"=B{r}*I{r}")
        ws_mp.cell(row=r, column=11, value=f"=C{r}*I{r}")
        ws_mp.cell(row=r, column=12, value=f"=D{r}*I{r}")
        ws_mp.cell(row=r, column=13, value=item[6])
        ws_mp.cell(row=r, column=14, value=f"=B{r}*M{r}")
        ws_mp.cell(row=r, column=15, value=f"=C{r}*M{r}")
        ws_mp.cell(row=r, column=16, value=f"=D{r}*M{r}")
        ws_mp.cell(row=r, column=17, value=item[7])
        
        fill = fill_allergy if item[8] else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))
        font = font_allergy if item[8] else font_regular
        for c in range(1, 18):
            cell = ws_mp.cell(row=r, column=c)
            cell.border = border_all
            cell.font = font
            if fill.fill_type:
                cell.fill = fill
            if c == 1 or c == 17:
                cell.alignment = align_left
            elif c in [2, 3, 4, 5, 13]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
            if c in [6, 7, 8]:
                cell.number_format = '#,##0" kcal"'
            elif c in [9, 10, 11, 12]:
                cell.number_format = '"$"#,##0.00'
            elif c in [14, 15, 16]:
                cell.number_format = '0.0" oz"'
                
    mon_lunch_end = r
    
    # Mon Lunch Total
    r += 1
    ws_mp.cell(row=r, column=1, value="Lunch Total")
    ws_mp.cell(row=r, column=6, value=f"=SUM(F{mon_lunch_start}:F{mon_lunch_end})")
    ws_mp.cell(row=r, column=7, value=f"=SUM(G{mon_lunch_start}:G{mon_lunch_end})")
    ws_mp.cell(row=r, column=8, value=f"=SUM(H{mon_lunch_start}:H{mon_lunch_end})")
    ws_mp.cell(row=r, column=10, value=f"=SUM(J{mon_lunch_start}:J{mon_lunch_end})")
    ws_mp.cell(row=r, column=11, value=f"=SUM(K{mon_lunch_start}:K{mon_lunch_end})")
    ws_mp.cell(row=r, column=12, value=f"=SUM(L{mon_lunch_start}:L{mon_lunch_end})")
    ws_mp.cell(row=r, column=14, value=f"=SUM(N{mon_lunch_start}:N{mon_lunch_end})")
    ws_mp.cell(row=r, column=15, value=f"=SUM(O{mon_lunch_start}:O{mon_lunch_end})")
    ws_mp.cell(row=r, column=16, value=f"=SUM(P{mon_lunch_start}:P{mon_lunch_end})")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    mon_lunch_total_row = r
    
    # MONDAY DAILY TOTAL
    r += 1
    ws_mp.cell(row=r, column=1, value="MON APR 6 — DAILY TOTAL (Per Bear Can)")
    ws_mp.cell(row=r, column=6, value=f"=F{mon_bkfst_total_row}+F{mon_snacks_total_row}+F{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=7, value=f"=G{mon_bkfst_total_row}+G{mon_snacks_total_row}+G{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=8, value=f"=H{mon_bkfst_total_row}+H{mon_snacks_total_row}+H{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=10, value=f"=J{mon_bkfst_total_row}+J{mon_snacks_total_row}+J{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=11, value=f"=K{mon_bkfst_total_row}+K{mon_snacks_total_row}+K{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=12, value=f"=L{mon_bkfst_total_row}+L{mon_snacks_total_row}+L{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=14, value=f"=N{mon_bkfst_total_row}+N{mon_snacks_total_row}+N{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=15, value=f"=O{mon_bkfst_total_row}+O{mon_snacks_total_row}+O{mon_lunch_total_row}")
    ws_mp.cell(row=r, column=16, value=f"=P{mon_bkfst_total_row}+P{mon_snacks_total_row}+P{mon_lunch_total_row}")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_navy
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.border = border_grand_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    mon_daily_total_row = r
    
    # MONDAY PER PERSON
    r += 1
    ws_mp.cell(row=r, column=1, value="MON APR 6 — CALORIES DELIVERED PER PERSON")
    ws_mp.cell(row=r, column=6, value=f"=F{mon_daily_total_row}/3") # Scout 3
    ws_mp.cell(row=r, column=7, value=f"=G{mon_daily_total_row}/4") # Scout 4
    ws_mp.cell(row=r, column=8, value=f"=H{mon_daily_total_row}/3") # Adult 3
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = Border(top=thin_side, bottom=thick_top)
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
            
    mon_per_person_row = r
    
    # --- GRAND TOTALS FOR ALL 3 DAYS ---
    r += 1
    write_subheader(ws_mp, r, "GRAND TOTALS — COMBINED 3 DAYS", fill_navy)
    r += 1
    ws_mp.cell(row=r, column=1, value="TOTAL CALORIES & COST PER BEAR CAN TYPE")
    ws_mp.cell(row=r, column=6, value=f"=F{sat_daily_total_row}+F{sun_daily_total_row}+F{mon_daily_total_row}")
    ws_mp.cell(row=r, column=7, value=f"=G{sat_daily_total_row}+G{sun_daily_total_row}+G{mon_daily_total_row}")
    ws_mp.cell(row=r, column=8, value=f"=H{sat_daily_total_row}+H{sun_daily_total_row}+H{mon_daily_total_row}")
    ws_mp.cell(row=r, column=10, value=f"=J{sat_daily_total_row}+J{sun_daily_total_row}+J{mon_daily_total_row}")
    ws_mp.cell(row=r, column=11, value=f"=K{sat_daily_total_row}+K{sun_daily_total_row}+K{mon_daily_total_row}")
    ws_mp.cell(row=r, column=12, value=f"=L{sat_daily_total_row}+L{sun_daily_total_row}+L{mon_daily_total_row}")
    ws_mp.cell(row=r, column=14, value=f"=N{sat_daily_total_row}+N{sun_daily_total_row}+N{mon_daily_total_row}")
    ws_mp.cell(row=r, column=15, value=f"=O{sat_daily_total_row}+O{sun_daily_total_row}+O{mon_daily_total_row}")
    ws_mp.cell(row=r, column=16, value=f"=P{sat_daily_total_row}+P{sun_daily_total_row}+P{mon_daily_total_row}")
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_grand_total
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
        elif c in [10, 11, 12]:
            cell.number_format = '"$"#,##0.00'
        elif c in [14, 15, 16]:
            cell.number_format = '0.0" oz"'
            
    grand_total_row = r
    
    # GRAND TOTAL PER PERSON
    r += 1
    ws_mp.cell(row=r, column=1, value="TRIP TOTAL CALORIES DELIVERED PER PERSON")
    ws_mp.cell(row=r, column=6, value=f"=F{grand_total_row}/3") # Scout 3
    ws_mp.cell(row=r, column=7, value=f"=G{grand_total_row}/4") # Scout 4
    ws_mp.cell(row=r, column=8, value=f"=H{grand_total_row}/3") # Adult 3
    for c in range(1, 18):
        cell = ws_mp.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = Border(top=thin_side, bottom=thick_top)
        if c in [6, 7, 8]:
            cell.number_format = '#,##0" kcal"'
            
    trip_per_person_row = r
    
    
    # --- 3. PACKING LIST SHEET ---
    ws_pl = wb.create_sheet(title="Packing List")
    ws_pl.views.sheetView[0].showGridLines = True
    
    ws_pl["A1"] = "OHLONE WILDERNESS TRAIL  ·  Bear Can Packing List  ·  April 4–6, 2026"
    ws_pl["A1"].font = font_title
    ws_pl["A2"] = "Check off each item as you pack.   ★ = adult cans only.   Qty = items per bear can type."
    ws_pl["A2"].font = font_subtitle
    
    pl_headers = [
        "Packed", "Meal", "Item", "Scout Can\n(3 ppl) Qty", "Scout Can\n(4 ppl) Qty", "Adult Can\n(3 ppl) Qty", "Notes"
    ]
    
    # Set headers
    for col_idx, h in enumerate(pl_headers, start=1):
        cell = ws_pl.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_wrap_center
        cell.border = border_all
    ws_pl.row_dimensions[4].height = 28
    
    # We will build packing list by parsing our meal plan data
    # Saturday
    pl_rows = [
        ("SATURDAY  APR 4", "", "", "", "", "", "", True),
        ("", "SNACKS", "Kind PB Dark Choc Bar", 3, 4, 3, "hip-belt pocket; ⚠ NUT", False),
        ("", "SNACKS", "Kind – Caramel Almond & Peanut Bar", 3, 4, 3, "hip-belt pocket; ⚠ NUT", False),
        ("", "SNACKS", "Krave Sea Salt Beef Jerky", 3, 4, 3, "hip-belt pocket; nut-free", False),
        ("", "SNACKS", "Gatorade Powder Pack", 3, 4, 3, "hip-belt pocket; electrolytes", False),
        ("", "SNACKS", "Trader Joe's Dried Mango Slices (1 oz)", 3, 4, 3, "hip-belt pocket; nut-free", False),
        ("", "LUNCH", "Columbus Hard Salami (packet)", 3, 4, 3, "no cook; nut-free", False),
        ("", "LUNCH", "Wasa Crispbread", 3, 4, 3, "no cook; nut-free", False),
        ("", "LUNCH", "Tillamook Cheddar Cheese Packets", 3, 4, 3, "no cook; nut-free", False),
        ("", "DINNER", "Peak Refuel – Beef Pasta Marinara", 2, 3, 3, "hot · 1.5 svgs/pouch", False),
        ("", "DINNER", "Flour Tortillas", 6, 8, 6, "dinner wraps + side", False),
        ("", "DINNER", "Tillamook Cheddar Cheese Packets", 6, 8, 6, "dinner side", False),
        ("", "DINNER", "Swiss Miss Hot Chocolate", 6, 8, 0, "scouts only", False),
        ("", "DESSERT", "Backpacker's Pantry Crème Brûlée", 2, 2, 2, "hot · buddy-pair shares 1", False),
        
        ("SUNDAY  APR 5", "", "", "", "", "", "", True),
        ("", "BREAKFAST", "MET-Rx Big100 – Super Cookie Crunch", 3, 4, 3, "grab & go; ⚠ NUT", False),
        ("", "BREAKFAST", "Swiss Miss Hot Chocolate", 6, 8, 0, "scouts only", False),
        ("", "BREAKFAST", "Adults: own coffee / tea", 0, 0, 3, "adults only ★", False),
        ("", "SNACKS", "Krave Sea Salt Original Beef Jerky", 3, 4, 3, "hip-belt pocket; nut-free", False),
        ("", "SNACKS", "Kind – Dark Choc Nuts & Sea Salt Bar", 3, 4, 3, "hip-belt pocket; ⚠ NUT", False),
        ("", "SNACKS", "Gatorade Powder Pack", 6, 8, 6, "×2/person; climb day", False),
        ("", "SNACKS", "Banana Chips (1 oz)", 3, 4, 3, "nut-free", False),
        ("", "LUNCH", "Chicken of the Sea Shredded Chicken Pouch", 3, 4, 3, "no cook; nut-free", False),
        ("", "LUNCH", "Flour Tortillas", 6, 8, 6, "lunch wraps", False),
        ("", "LUNCH", "Tillamook Cheddar Cheese Packets", 3, 4, 3, "lunch side", False),
        ("", "LUNCH", "Clif Bar – Crunchy PB", 3, 4, 3, "lunch snack; ⚠ NUT", False),
        ("", "LUNCH", "Justin's Classic PB Squeeze Pack", 3, 4, 3, "lunch side; ⚠ NUT", False),
        ("", "SUPPLEMENT", "Clif Builder's Bar (adult supplement)", 0, 0, 3, "pre-dinner; adults only ★; ⚠ NUT", False),
        ("", "SUPPLEMENT", "Justin's PB Squeeze (adult supplement)", 0, 0, 3, "any time; adults only ★; ⚠ NUT", False),
        ("", "DINNER", "Peak Refuel – Beef Stroganoff", 2, 3, 3, "hot · 1.5 svgs/pouch", False),
        ("", "DINNER", "Flour Tortillas", 6, 8, 6, "dinner wraps", False),
        ("", "DINNER", "Idahoan Mashed Potatoes", 3, 4, 3, "dinner side", False),
        ("", "DINNER", "Swiss Miss Hot Chocolate", 6, 8, 0, "scouts only", False),
        ("", "DESSERT", "Justin's Dark Choc PB Cups", 3, 4, 3, "1 packet/person; ⚠ NUT", False),
        
        ("MONDAY  APR 6", "", "", "", "", "", "", True),
        ("", "BREAKFAST", "ProBar Meal – Oatmeal Chocolate Chip", 3, 4, 3, "grab & go; ⚠ NUT (cashews)", False),
        ("", "BREAKFAST", "Swiss Miss Hot Chocolate", 6, 8, 0, "scouts only", False),
        ("", "SNACKS", "Clif Bar – Crunchy PB", 3, 4, 3, "hip-belt pocket; ⚠ NUT", False),
        ("", "SNACKS", "Krave Sea Salt Original Beef Jerky", 3, 4, 3, "hip-belt pocket; nut-free", False),
        ("", "SNACKS", "Gatorade Powder Pack", 3, 4, 3, "hip-belt pocket; electrolytes", False),
        ("", "SNACKS", "Peak Refuel – Fudge Brownie Bites", 2, 2, 2, "Mon afternoon snack; nut-free", False),
        ("", "LUNCH", "Flour Tortillas", 6, 8, 6, "no cook; lunch wraps", False),
        ("", "LUNCH", "Justin's Classic PB Squeeze Pack", 3, 4, 3, "lunch wraps; ⚠ NUT", False)
    ]
    
    pl_row = 5
    for item in pl_rows:
        is_day = item[7]
        if is_day:
            # Day divider row
            ws_pl.merge_cells(start_row=pl_row, start_column=1, end_row=pl_row, end_column=7)
            cell = ws_pl.cell(row=pl_row, column=1, value=item[0])
            cell.font = font_subheader
            cell.fill = fill_teal
            cell.alignment = align_left
            ws_pl.row_dimensions[pl_row].height = 22
        else:
            # Regular packing list item
            ws_pl.cell(row=pl_row, column=1, value="☐").alignment = align_center
            ws_pl.cell(row=pl_row, column=2, value=item[1]).alignment = align_left
            ws_pl.cell(row=pl_row, column=3, value=item[2]).alignment = align_left
            ws_pl.cell(row=pl_row, column=4, value=item[3]).alignment = align_center
            ws_pl.cell(row=pl_row, column=5, value=item[4]).alignment = align_center
            ws_pl.cell(row=pl_row, column=6, value=item[5]).alignment = align_center
            ws_pl.cell(row=pl_row, column=7, value=item[6]).alignment = align_left
            
            # Formatting
            is_nut = "⚠ NUT" in item[6]
            is_star = "★" in item[6]
            
            fill = fill_allergy if is_nut else (fill_adult if is_star else (fill_alt_row if pl_row % 2 == 1 else PatternFill(fill_type=None)))
            font = font_allergy if is_nut else (font_adult if is_star else font_regular)
            
            for c in range(1, 8):
                cell = ws_pl.cell(row=pl_row, column=c)
                cell.font = font
                cell.border = border_all
                if fill.fill_type:
                    cell.fill = fill
                    
            ws_pl.cell(row=pl_row, column=1).font = Font(name="Segoe UI", size=10, bold=True)
            
        pl_row += 1
        
    # --- 4. SHOPPING LIST SHEET ---
    ws_sl = wb.create_sheet(title="Shopping List")
    ws_sl.views.sheetView[0].showGridLines = True
    
    ws_sl["A1"] = "Master Shopping List — 19 People · 6 Bear Cans  |  April 4–6, 2026"
    ws_sl["A1"].font = font_title
    ws_sl["A2"] = "🟡 Yellow = bulk buy (Costco/Amazon)  |  🟠 Orange = adult supplement  |  Group Total = 3*(Scout 3) + 1*(Scout 4) + 2*(Adult 3)"
    ws_sl["A2"].font = font_subtitle
    
    sl_headers = [
        "☐", "Food Item", "Qty\n(Scout 3 Can)", "Qty\n(Scout 4 Can)", "Qty\n(Adult 3 Can)", 
        "Group Total\n(Units/Packs)", "$/Unit", "Scout 3\nCan $", "Scout 4\nCan $", "Adult 3\nCan $",
        "Group Total\n$", "oz/unit", "Group\nozs", "Where to Buy"
    ]
    
    for col_idx, h in enumerate(sl_headers, start=1):
        cell = ws_sl.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_wrap_center
        cell.border = border_all
    ws_sl.row_dimensions[3].height = 30
    
    # Shopping list data
    # item_name, sc3, sc4, ad3, price, oz, notes, is_nut, is_supp, is_bulk
    sl_data = [
        ["MET-Rx Big100 – Super Cookie Crunch", 3, 4, 3, 2.19, 3.53, "⚠ NUT ALLERGY SWAP: No Nuts! Chocolate Chip bar", True, False, True],
        ["ProBar Meal – Oatmeal Chocolate Chip", 3, 4, 3, 2.99, 3.00, "⚠ NUT ALLERGY SWAP: No Nuts! bar", True, False, True],
        ["Peak Refuel – Beef Pasta Marinara", 2, 3, 3, 7.49, 3.20, "Sat dinner (1.5 svgs/pouch)", False, False, False],
        ["Peak Refuel – Beef Stroganoff", 2, 3, 3, 7.49, 3.50, "Sun dinner (1.5 svgs/pouch)", False, False, False],
        ["Backpacker's Pantry – Crème Brûlée", 2, 2, 2, 8.99, 2.26, "Sat dessert (buddy-pair shares)", False, False, False],
        ["Justin's Dark Chocolate Peanut Butter Cups", 3, 4, 3, 1.49, 1.41, "⚠ NUT ALLERGY SWAP: Honey Stinger waffle", True, False, False],
        ["Columbus Hard Salami (sliced packet)", 3, 4, 3, 2.99, 2.00, "Sat lunch; 1/person", False, False, True],
        ["Wasa Crisp'n Light Crackerbread", 3, 4, 3, 0.50, 1.00, "Sat lunch; 1/person", False, False, False],
        ["Chicken of the Sea Shredded Chicken Pouch", 3, 4, 3, 2.29, 2.50, "Sun lunch; 1/person", False, False, True],
        ["Mission Street Taco Flour Tortillas", 18, 24, 18, 0.25, 0.50, "6/person (Sat din + Sun lunch + Mon lunch)", False, False, True],
        ["Tillamook Cheddar Cheese Packets", 12, 16, 12, 0.75, 0.70, "4/person (Sat din ×2 + Sat lunch ×1 + Sun lunch ×1)", False, False, True],
        ["Justin's Classic PB Squeeze Pack (lunch)", 6, 8, 6, 1.49, 1.15, "Sun lunch + Mon lunch; ⚠ NUT", True, False, False],
        ["Idahoan Mashed Potatoes – Butter & Herb", 3, 4, 3, 0.98, 1.00, "Sun dinner; 1/person", False, False, False],
        ["Kind – Peanut Butter Dark Chocolate Bar", 3, 4, 3, 1.04, 1.40, "Sat snacks only; ⚠ NUT", True, False, True],
        ["Kind – Caramel Almond Bar (Saturday snacks)", 3, 4, 3, 1.59, 1.40, "Sat snacks only; ⚠ NUT", True, False, False],
        ["Kind – Dark Choc Nuts & Sea Salt Bar", 3, 4, 3, 1.04, 1.40, "Sun snacks only; ⚠ NUT", True, False, True],
        ["Krave Sea Salt Original Beef Jerky", 9, 12, 9, 2.19, 1.48, "1/person × 3 snack days", False, False, True],
        ["Clif Bar – Crunchy Peanut Butter", 6, 8, 6, 1.29, 2.40, "Sun lunch + Mon snack; ⚠ NUT", True, False, True],
        ["Clif Builder's – Crunchy Peanut Butter", 0, 0, 3, 2.17, 2.40, "Adult supplement; ⚠ NUT", True, True, True],
        ["Justin's Classic PB Squeeze Pack (adult supp)", 0, 0, 3, 1.49, 1.15, "Adult supplement; ⚠ NUT", True, True, False],
        ["Swiss Miss Hot Chocolate", 12, 16, 0, 0.38, 0.73, "Scout cans only; Sat din + Sun bkfst + Sun din + Mon bkfst", False, False, True],
        ["Gatorade Powder Pack", 12, 16, 12, 0.38, 1.23, "4/person (1+2+1 days)", False, False, True],
        ["Peak Refuel – Fudge Brownie Bites", 2, 2, 2, 4.99, 1.50, "Mon afternoon snack · no cook", False, False, False],
        ["Trader Joe's Dried Mango Slices (1 oz/person)", 3, 4, 3, 0.55, 1.00, "Sat snack · nut-free ✓", False, False, False],
        ["Banana Chips (1 oz/person)", 3, 4, 3, 0.45, 1.00, "Sun snack · nut-free ✓", False, False, True],
        
        # Nut allergy swaps
        ["88 Acres Seed + Oat Bars (variety — allergy scout)", 0, 0, 0, 2.40, 1.40, "Allergy scout: replaces Kind PB (Sat) + Kind DCS (Sun)", True, False, False],
        ["Once Again Organic Sunflower Butter Squeeze Pack", 0, 0, 0, 1.50, 1.15, "Allergy scout: replaces Justin's PB on Sun + Mon lunch", True, False, False],
        ["Honey Stinger Gold Waffle (allergy scout contingency)", 0, 0, 0, 1.09, 1.00, "Allergy scout: replaces Crème Brûlée or PB Cups", True, False, False]
    ]
    
    # Note: for allergy swaps, they are pre-packed. The quantities in canisters are 0, but group totals are:
    # 88 Acres: 2 packs (replaces Kind PB on Sat, Kind DCS on Sun) -> wait, actually 3 packs: Sat snacks ×2 (PB & Caramel), Sun snacks ×1 (Kind DCS) = 3 packs!
    # Once Again Sunflower Squeeze: 3 packs (replaces Justin's PB on Sun lunch, Mon lunch, and Sun adult supp? Wait, the allergy scout is a scout, not an adult, so no adult supplement needed. So they need it for Sun lunch and Mon lunch = 2 packs!).
    # Honey Stinger waffle: 2 packs (replaces BP Crème Brûlée Sat dessert, Justin's PB Cups Sun dessert = 2 packs!).
    # Let's adjust the hardcoded Group Totals for allergy swaps:
    # Row 27 (88 Acres): 3 units.
    # Row 28 (Once Again): 2 units.
    # Row 29 (Honey Stinger): 2 units.
    
    sl_start_row = 4
    for idx, item in enumerate(sl_data):
        r = sl_start_row + idx
        ws_sl.cell(row=r, column=1, value="☐").alignment = align_center
        ws_sl.cell(row=r, column=2, value=item[0]).alignment = align_left
        ws_sl.cell(row=r, column=3, value=item[1]).alignment = align_center
        ws_sl.cell(row=r, column=4, value=item[2]).alignment = align_center
        ws_sl.cell(row=r, column=5, value=item[3]).alignment = align_center
        
        # Group Total formula: 3 * Can1-3 + 1 * Can4 + 2 * Can5-6
        # Except for allergy swap items where we set group totals manually!
        is_allergy_item = "allergy scout" in item[6].lower() or "88 acres" in item[0].lower() or "once again" in item[0].lower() or "honey stinger" in item[0].lower()
        if is_allergy_item:
            if "88 acres" in item[0].lower():
                ws_sl.cell(row=r, column=6, value=3)
            elif "once again" in item[0].lower():
                ws_sl.cell(row=r, column=6, value=2)
            else: # Honey stinger
                ws_sl.cell(row=r, column=6, value=2)
        else:
            ws_sl.cell(row=r, column=6, value=f"=3*C{r}+1*D{r}+2*E{r}")
            
        ws_sl.cell(row=r, column=6).alignment = align_center
        
        # $/Unit
        ws_sl.cell(row=r, column=7, value=item[4]).alignment = align_right
        ws_sl.cell(row=r, column=7).number_format = '"$"#,##0.00'
        
        # Cost columns (H, I, J, K) - CORRECTED formulas (same-row references!)
        ws_sl.cell(row=r, column=8, value=f"=C{r}*G{r}").alignment = align_right
        ws_sl.cell(row=r, column=8).number_format = '"$"#,##0.00'
        ws_sl.cell(row=r, column=9, value=f"=D{r}*G{r}").alignment = align_right
        ws_sl.cell(row=r, column=9).number_format = '"$"#,##0.00'
        ws_sl.cell(row=r, column=10, value=f"=E{r}*G{r}").alignment = align_right
        ws_sl.cell(row=r, column=10).number_format = '"$"#,##0.00'
        ws_sl.cell(row=r, column=11, value=f"=F{r}*G{r}").alignment = align_right
        ws_sl.cell(row=r, column=11).number_format = '"$"#,##0.00'
        
        # Weight columns (L, M)
        ws_sl.cell(row=r, column=12, value=item[5]).alignment = align_center
        ws_sl.cell(row=r, column=12).number_format = '0.00" oz"'
        ws_sl.cell(row=r, column=13, value=f"=F{r}*L{r}").alignment = align_right
        ws_sl.cell(row=r, column=13).number_format = '0.0" oz"'
        
        # Where to buy
        ws_sl.cell(row=r, column=14, value=item[6]).alignment = align_left
        
        # Formatting styles
        is_nut = item[7]
        is_supp = item[8]
        is_bulk = item[9]
        
        fill = fill_allergy if is_nut else (fill_adult if is_supp else (fill_light_gray if is_bulk else (fill_alt_row if r % 2 == 1 else PatternFill(fill_type=None))))
        font = font_allergy if is_nut else (font_adult if is_supp else font_regular)
        
        for c in range(1, 15):
            cell = ws_sl.cell(row=r, column=c)
            cell.font = font
            cell.border = border_all
            if fill.fill_type:
                cell.fill = fill
                
        ws_sl.cell(row=r, column=1).font = Font(name="Segoe UI", size=10, bold=True)
        
    sl_end_row = sl_start_row + len(sl_data) - 1
    
    # Grand Total row
    r = sl_end_row + 1
    ws_sl.cell(row=r, column=1, value="")
    ws_sl.cell(row=r, column=2, value="GRAND TOTAL — ALL FOOD (19 people)").font = font_bold
    ws_sl.cell(row=r, column=8, value=f"=SUM(H{sl_start_row}:H{sl_end_row})").number_format = '"$"#,##0.00'
    ws_sl.cell(row=r, column=9, value=f"=SUM(I{sl_start_row}:I{sl_end_row})").number_format = '"$"#,##0.00'
    ws_sl.cell(row=r, column=10, value=f"=SUM(J{sl_start_row}:J{sl_end_row})").number_format = '"$"#,##0.00'
    ws_sl.cell(row=r, column=11, value=f"=SUM(K{sl_start_row}:K{sl_end_row})").number_format = '"$"#,##0.00'
    ws_sl.cell(row=r, column=13, value=f"=SUM(M{sl_start_row}:M{sl_end_row})").number_format = '0.0" oz"'
    
    for c in range(1, 15):
        cell = ws_sl.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_grand_total
        if c in [8, 9, 10, 11]:
            cell.alignment = align_right
        elif c == 13:
            cell.alignment = align_right
            
    # Per person total row
    r += 1
    ws_sl.cell(row=r, column=2, value="Per Person Cost (Group Total ÷ 19)").font = font_bold
    ws_sl.cell(row=r, column=11, value=f"=K{r-1}/19").number_format = '"$"#,##0.00'
    for c in range(1, 15):
        cell = ws_sl.cell(row=r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = border_section_total
        if c == 11:
            cell.alignment = align_right
            
            
    # --- 5. NUT ALLERGY — SCOUT PACK SHEET ---
    ws_na = wb.create_sheet(title="Nut Allergy — Scout Pack")
    ws_na.views.sheetView[0].showGridLines = True
    
    ws_na["A1"] = "Nut Allergy Accommodation — Individual Scout Carry Pack"
    ws_na["A1"].font = font_title
    ws_na["A2"] = "Pre-pack all items in a labelled zip-lock or dry bag for this scout. They carry it personally and swap at each nut-item moment. Main canisters are UNCHANGED."
    ws_na["A2"].font = font_subtitle
    
    na_headers = [
        "Day", "Meal", "Remove (has nuts)", "Nuts in it", "Replace with", "Why safe", "kcal", "Where to buy"
    ]
    
    for col_idx, h in enumerate(na_headers, start=1):
        cell = ws_na.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all
    ws_na.row_dimensions[4].height = 24
    
    na_data = [
        ["Sat Apr 4", "Snacks", "Kind PB Dark Choc Bar", "Peanuts, almonds", "88 Acres Dark Choc Sea Salt Seed + Oat Bar", "Dedicated nut-free bakery. Free from top 9.", 150, "Amazon / 88acres.com"],
        ["Sat Apr 4", "Snacks", "Kind – Caramel Almond & Peanut Bar", "Almonds, peanuts", "88 Acres Cinnamon Maple Seed + Oat Bar", "Same nut-free facility.", 150, "Amazon / 88acres.com"],
        ["Sat Apr 4", "Dessert", "BP Crème Brûlée", "Check label - may contain cashews", "Honey Stinger Gold Waffle (contingency)", "Honey Stinger: simple ingredients, no nuts.", 160, "REI / Amazon"],
        ["Sun Apr 5", "Breakfast", "MET-Rx Big100 Super Cookie Crunch", "Peanuts, almonds, tree nuts", "No Nuts! Chocolate Chip Protein Bar", "Certified 100% nut-free facility. 12g protein.", 185, "Amazon / gononuts.com"],
        ["Sun Apr 5", "Snacks", "Kind – Dark Choc Nuts & Sea Salt Bar", "Almonds, peanuts, cashews", "88 Acres bar (any flavor)", "Same nut-free facility.", 150, "Amazon / 88acres.com"],
        ["Sun Apr 5", "Lunch", "Clif Bar – Crunchy PB", "Peanuts", "Clif Bar – Chocolate Chip", "Clif Chocolate Chip: no nuts in ingredients.", 250, "Add to main Costco order"],
        ["Sun Apr 5", "Lunch", "Justin's Classic PB Squeeze Pack", "Peanuts", "Once Again Sunflower Butter Squeeze Pack", "Peanut-free facility. Free from top 9.", 200, "Amazon"],
        ["Sun Apr 5", "Dessert", "Justin's Dark Choc PB Cups", "Peanuts, almonds", "Honey Stinger Gold Waffle or Enjoy Life Choc", "No nuts. Enjoy Life: top-14 free.", 160, "REI / Amazon"],
        ["Mon Apr 6", "Breakfast", "ProBar Meal – Oatmeal Chocolate Chip", "Cashews", "No Nuts! Cinnamon Roll or Caramel Mocha", "Same certified nut-free facility.", 185, "Amazon / gononuts.com"],
        ["Mon Apr 6", "Snacks", "Clif Bar – Crunchy PB", "Peanuts", "Clif Bar – Chocolate Chip", "Clif Chocolate Chip: no nuts in ingredients.", 250, "Add to main Costco order"],
        ["Mon Apr 6", "Lunch", "Justin's Classic PB Squeeze Pack", "Peanuts", "Once Again Sunflower Butter Squeeze Pack", "Peanut-free facility. Free from top 9.", 200, "Amazon"]
    ]
    
    for idx, item in enumerate(na_data):
        r = 5 + idx
        for c in range(1, 9):
            cell = ws_na.cell(row=r, column=c, value=item[c-1])
            cell.font = font_regular
            cell.border = border_all
            if r % 2 == 1:
                cell.fill = fill_alt_row
            
            if c in [1, 2, 7]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if c == 7:
                cell.number_format = '#,##0" kcal"'
                
    # --- AUTO-FIT COLUMN WIDTHS & SET ALIGNMENTS ---
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if cell.coordinate in ws.merged_cells:
                    # Ignore title rows for column widths
                    continue
                if '\n' in val:
                    lines = val.split('\n')
                    max_len = max(max_len, max(len(l) for l in lines))
                else:
                    max_len = max(max_len, len(val))
            
            col_letter = get_column_letter(col[0].column)
            # Apply padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 10)
            
    # Explicitly set some specific column widths for better aesthetic look
    ws_ov.column_dimensions['A'].width = 24
    ws_ov.column_dimensions['C'].width = 15
    ws_ov.column_dimensions['D'].width = 15
    ws_ov.column_dimensions['E'].width = 15
    ws_ov.column_dimensions['F'].width = 15
    ws_ov.column_dimensions['G'].width = 15
    ws_ov.column_dimensions['J'].width = 32
    
    ws_mp.column_dimensions['A'].width = 36
    ws_mp.column_dimensions['Q'].width = 45
    
    ws_pl.column_dimensions['C'].width = 36
    ws_pl.column_dimensions['G'].width = 40
    
    ws_sl.column_dimensions['B'].width = 36
    ws_sl.column_dimensions['N'].width = 45
    
    ws_na.column_dimensions['C'].width = 28
    ws_na.column_dimensions['E'].width = 32
    ws_na.column_dimensions['F'].width = 32
    ws_na.column_dimensions['H'].width = 24

    wb.save("spring_hike_menu_19_people.xlsx")
    print("Workbook generated successfully as spring_hike_menu_19_people.xlsx!")

if __name__ == "__main__":
    create_spring_hike_menu()

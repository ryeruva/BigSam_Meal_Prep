import openpyxl

def run_verification():
    wb = openpyxl.load_workbook("spring_hike_menu_19_people.xlsx", data_only=False)
    
    # 1. Verify Overview sheet canister assignment and counts
    ws_ov = wb["Overview"]
    can1_ppl = ws_ov["B7"].value
    can2_ppl = ws_ov["C7"].value
    can3_ppl = ws_ov["D7"].value
    can4_ppl = ws_ov["E7"].value
    can5_ppl = ws_ov["F7"].value
    can6_ppl = ws_ov["G7"].value
    
    print("Overview Canister Configuration:")
    print(f"Can 1: {can1_ppl}")
    print(f"Can 2: {can2_ppl}")
    print(f"Can 3: {can3_ppl}")
    print(f"Can 4: {can4_ppl}")
    print(f"Can 5: {can5_ppl}")
    print(f"Can 6: {can6_ppl}")
    
    # 2. Check Shopping List sheet formulas and row alignments
    ws_sl = wb["Shopping List"]
    errors = 0
    
    # Dynamically find the end row of the shopping list items
    sl_end_row = None
    for r in range(4, ws_sl.max_row + 1):
        val = ws_sl.cell(row=r, column=2).value
        if val == "Grand Total":
            sl_end_row = r - 1
            break
    if sl_end_row is None:
        print("Error: Could not find 'Grand Total' row in Shopping List")
        sl_end_row = 31 # fallback
        
    print(f"Detected Shopping List items end at row {sl_end_row}")

    # Check rows 4 to sl_end_row
    for r in range(4, sl_end_row + 1):
        item_name = ws_sl.cell(row=r, column=2).value
        qty_p1_sc3 = ws_sl.cell(row=r, column=3).value or 0
        qty_p1_sc4 = ws_sl.cell(row=r, column=4).value or 0
        qty_p1_ad3 = ws_sl.cell(row=r, column=5).value or 0
        qty_p2_sc3 = ws_sl.cell(row=r, column=6).value or 0
        qty_p2_sc4 = ws_sl.cell(row=r, column=7).value or 0
        qty_p2_ad3 = ws_sl.cell(row=r, column=8).value or 0
        group_total = ws_sl.cell(row=r, column=9).value
        price = ws_sl.cell(row=r, column=10).value
        cost_sc3 = ws_sl.cell(row=r, column=11).value
        cost_sc4 = ws_sl.cell(row=r, column=12).value
        cost_ad3 = ws_sl.cell(row=r, column=13).value
        cost_group = ws_sl.cell(row=r, column=14).value
        weight_oz = ws_sl.cell(row=r, column=15).value
        weight_group = ws_sl.cell(row=r, column=16).value
        notes = ws_sl.cell(row=r, column=17).value or ""
        
        # Verify Group Total formula
        is_allergy_item = "allergy scout" in notes.lower() or "88 acres" in item_name.lower() or "once again" in item_name.lower() or "honey stinger" in item_name.lower()
        if not is_allergy_item:
            expected_group_total_formula = f"=3*(C{r}+F{r})+1*(D{r}+G{r})+2*(E{r}+H{r})"
            if group_total != expected_group_total_formula:
                print(f"Row {r} Error: Group total formula is {group_total}, expected {expected_group_total_formula}")
                errors += 1
                
        # Verify Cost Formulas row alignment
        expected_cost_sc3 = f"=(C{r}+F{r})*J{r}"
        expected_cost_sc4 = f"=(D{r}+G{r})*J{r}"
        expected_cost_ad3 = f"=(E{r}+H{r})*J{r}"
        expected_cost_group = f"=I{r}*J{r}"
        
        if cost_sc3 != expected_cost_sc3:
            print(f"Row {r} Error: Scout 3 Cost formula is {cost_sc3}, expected {expected_cost_sc3}")
            errors += 1
        if cost_sc4 != expected_cost_sc4:
            print(f"Row {r} Error: Scout 4 Cost formula is {cost_sc4}, expected {expected_cost_sc4}")
            errors += 1
        if cost_ad3 != expected_cost_ad3:
            print(f"Row {r} Error: Adult 3 Cost formula is {cost_ad3}, expected {expected_cost_ad3}")
            errors += 1
        if cost_group != expected_cost_group:
            print(f"Row {r} Error: Group Cost formula is {cost_group}, expected {expected_cost_group}")
            errors += 1
            
        # Verify Weight Formula row alignment
        expected_weight_group = f"=I{r}*O{r}"
        if weight_group != expected_weight_group:
            print(f"Row {r} Error: Group Weight formula is {weight_group}, expected {expected_weight_group}")
            errors += 1

    # Check Budget Overview cells in Overview
    cost_adult_formula = ws_ov["B35"].value
    cost_sc3_formula = ws_ov["C35"].value
    cost_sc4_formula = ws_ov["D35"].value
    cost_total_formula = ws_ov["E35"].value
    
    print("\nOverview Budget Formulas:")
    print(f"Adult Can Cost Cell B35: {cost_adult_formula}")
    print(f"Scout 3-ppl Cost Cell C35: {cost_sc3_formula}")
    print(f"Scout 4-ppl Cost Cell D35: {cost_sc4_formula}")
    print(f"Group Total Cost Cell E35: {cost_total_formula}")
    
    expected_cost_adult = f"=SUM('Shopping List'!M$4:M${sl_end_row})"
    expected_cost_sc3 = f"=SUM('Shopping List'!K$4:K${sl_end_row})"
    expected_cost_sc4 = f"=SUM('Shopping List'!L$4:L${sl_end_row})"
    expected_cost_total = f"=SUM('Shopping List'!N$4:N${sl_end_row})"
    
    if cost_adult_formula != expected_cost_adult:
        print(f"Overview Error: B35 is {cost_adult_formula}, expected {expected_cost_adult}")
        errors += 1
    if cost_sc3_formula != expected_cost_sc3:
        print(f"Overview Error: C35 is {cost_sc3_formula}, expected {expected_cost_sc3}")
        errors += 1
    if cost_sc4_formula != expected_cost_sc4:
        print(f"Overview Error: D35 is {cost_sc4_formula}, expected {expected_cost_sc4}")
        errors += 1
    if cost_total_formula != expected_cost_total:
        print(f"Overview Error: E35 is {cost_total_formula}, expected {expected_cost_total}")
        errors += 1
        
    # Check Overview Max Carry weight formulas
    weight_adult_formula = ws_ov["B37"].value
    weight_sc3_formula = ws_ov["C37"].value
    weight_sc4_formula = ws_ov["D37"].value
    weight_total_formula = ws_ov["E37"].value
    
    print("\nOverview Max Carry Weight Formulas:")
    print(f"Adult Max Carry Weight Cell B37: {weight_adult_formula}")
    print(f"Scout 3-ppl Max Carry Weight Cell C37: {weight_sc3_formula}")
    print(f"Scout 4-ppl Max Carry Weight Cell D37: {weight_sc4_formula}")
    print(f"Group Max Carry Weight Cell E37: {weight_total_formula}")
    
    expected_weight_adult = f"=SUMPRODUCT('Shopping List'!E$4:E${sl_end_row},'Shopping List'!O$4:O${sl_end_row})/16"
    expected_weight_sc3 = f"=SUMPRODUCT('Shopping List'!C$4:C${sl_end_row},'Shopping List'!O$4:O${sl_end_row})/16"
    expected_weight_sc4 = f"=SUMPRODUCT('Shopping List'!D$4:D${sl_end_row},'Shopping List'!O$4:O${sl_end_row})/16"
    expected_weight_total = f"=SUMPRODUCT(3*'Shopping List'!C$4:C${sl_end_row}+1*'Shopping List'!D$4:D${sl_end_row}+2*'Shopping List'!E$4:E${sl_end_row},'Shopping List'!O$4:O${sl_end_row})/16"
    
    if weight_adult_formula != expected_weight_adult:
        print(f"Overview Error: B37 is {weight_adult_formula}, expected {expected_weight_adult}")
        errors += 1
    if weight_sc3_formula != expected_weight_sc3:
        print(f"Overview Error: C37 is {weight_sc3_formula}, expected {expected_weight_sc3}")
        errors += 1
    if weight_sc4_formula != expected_weight_sc4:
        print(f"Overview Error: D37 is {weight_sc4_formula}, expected {expected_weight_sc4}")
        errors += 1
    if weight_total_formula != expected_weight_total:
        print(f"Overview Error: E37 is {weight_total_formula}, expected {expected_weight_total}")
        errors += 1

    # 3. Check Meal Plan Formulas in Meal Plan (Per Bear Can) sheet
    ws_mp = wb["Meal Plan (Per Bear Can)"]
    print("\nMeal Plan Formula Previews:")
    
    item_count = 0
    # Let's dynamically find rows representing items
    for check_row in range(4, ws_mp.max_row + 1):
        item_name = ws_mp.cell(row=check_row, column=1).value
        # If item_name is empty or not a string, skip
        if not item_name or not isinstance(item_name, str):
            continue
        # If it is a header or subtotal or section description, skip
        if (item_name.startswith("  ") or 
            "TOTAL" in item_name or 
            "PHASE" in item_name or 
            "DAY" in item_name or 
            "BREAKFAST" in item_name or 
            "LUNCH" in item_name or 
            "DINNER" in item_name or 
            "SNACKS" in item_name or 
            "SUPPLEMENT" in item_name or 
            "DESSERT" in item_name or 
            "Food Item" in item_name or
            "Eaten at home" in item_name):
            continue
            
        sq3_qty = ws_mp.cell(row=check_row, column=2).value
        sq4_qty = ws_mp.cell(row=check_row, column=3).value
        ad3_qty = ws_mp.cell(row=check_row, column=4).value
        sq3_kcal = ws_mp.cell(row=check_row, column=6).value
        sq4_kcal = ws_mp.cell(row=check_row, column=7).value
        ad3_kcal = ws_mp.cell(row=check_row, column=8).value
        
        # We only print the first 5 previews
        if item_count < 5:
            print(f"Row {check_row} [{item_name}]: Qty({sq3_qty}, {sq4_qty}, {ad3_qty}) | kcal formulas({sq3_kcal}, {sq4_kcal}, {ad3_kcal})")
            
        item_count += 1
        
        # Check kcal calculations
        if sq3_kcal != f"=B{check_row}*E{check_row}":
            print(f"Row {check_row} Error: Scout 3 kcal is {sq3_kcal}, expected =B{check_row}*E{check_row}")
            errors += 1
        if sq4_kcal != f"=C{check_row}*E{check_row}":
            print(f"Row {check_row} Error: Scout 4 kcal is {sq4_kcal}, expected =C{check_row}*E{check_row}")
            errors += 1
        if ad3_kcal != f"=D{check_row}*E{check_row}":
            print(f"Row {check_row} Error: Adult 3 kcal is {ad3_kcal}, expected =D{check_row}*E{check_row}")
            errors += 1
            
    print(f"\nVerification finished. Total errors found: {errors}")
    if errors == 0:
        print("PASSED! Spreadsheet formulas are perfectly aligned and correct.")
    else:
        print("FAILED! Please correct the errors.")

if __name__ == "__main__":
    run_verification()
